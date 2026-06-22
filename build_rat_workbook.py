"""
Build an auditable XLSX twin of the HTML "Darcy Permeability (RAT)" calculator.

Design goal: AUDITABILITY for CalGEM. Every computed value is a live Excel
formula referencing named input cells, so a reviewer can click any result and
trace the math back to the governing equation and the source inputs. Field
units throughout (psi, ft, bbl/d, cP, mD), matching the HTML tool exactly.

Run:  python build_rat_workbook.py
Out:  Darcy-Permeability-RAT.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- styling ----
BRAND   = "1F4E5F"   # deep teal header
BAND    = "DCE6EA"   # section band
HILITE  = "FFF6E5"   # input cell tint
RESULT  = "EAF3EC"   # result tint
ANCHOR  = "1F4E5F"   # sensitivity base cell
POSCELL = "EAF3EC"   # within-50% sensitivity cell
GRID    = "B7C4CA"

f_title   = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
f_sub     = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
f_band    = Font(name="Calibri", size=11, bold=True, color="1F4E5F")
f_label   = Font(name="Calibri", size=10, bold=True, color="222222")
f_sym     = Font(name="Cambria Math", size=10, italic=True, color="444444")
f_val     = Font(name="Consolas", size=11, color="0B5A2E")
f_input   = Font(name="Consolas", size=11, bold=True, color="14406B")
f_units   = Font(name="Calibri", size=9, color="555555")
f_help    = Font(name="Calibri", size=9, italic=True, color="555555")
f_primary = Font(name="Consolas", size=20, bold=True, color="1F4E5F")
f_cite    = Font(name="Calibri", size=8.5, italic=True, color="666666")
f_eq      = Font(name="Consolas", size=9, color="333333")

thin = Side(style="thin", color=GRID)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

left   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
leftw  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
rightv = Alignment(horizontal="right",  vertical="center")

wb = Workbook()

# ============================================================== Calculator ===
ws = wb.active
ws.title = "Calculator"
ws.sheet_view.showGridLines = False
widths = {"A": 30, "B": 8, "C": 14, "D": 12, "E": 74}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

names = {}   # defined-name -> "Calculator!$C$r"

def band(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=text)
    c.font = f_band
    c.fill = PatternFill("solid", fgColor=BAND)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20

def help_height(text):
    # rough wrap estimate for column E (~74 chars wide)
    import math
    lines = max(1, math.ceil(len(text) / 92))
    return max(16, lines * 12 + 4)

def inp(ws, row, label, sym, value, units, helptext, name=None, is_input=True, numfmt="General"):
    ws.cell(row=row, column=1, value=label).font = f_label
    ws.cell(row=row, column=1).alignment = left
    ws.cell(row=row, column=2, value=sym).font = f_sym
    ws.cell(row=row, column=2).alignment = center
    c = ws.cell(row=row, column=3, value=value)
    c.font = f_input if is_input else f_val
    c.alignment = rightv
    c.border = box
    c.number_format = numfmt
    if is_input:
        c.fill = PatternFill("solid", fgColor=HILITE)
    ws.cell(row=row, column=4, value=units).font = f_units
    ws.cell(row=row, column=4).alignment = left
    h = ws.cell(row=row, column=5, value=helptext)
    h.font = f_help
    h.alignment = leftw
    ws.row_dimensions[row].height = help_height(helptext)
    if name:
        names[name] = f"Calculator!$C${row}"
    return row

r = 1
# --- title block ---
ws.merge_cells("A1:E1")
ws.cell(row=1, column=1, value="NSLLC Toolkit  —  Darcy Permeability (RAT)").font = f_title
ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=BRAND)
ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 26
ws.merge_cells("A2:E2")
ws.cell(row=2, column=1, value=("Effective (skin-included) permeability from a stabilized single-rate injectivity "
        "test (RAT survey) on a Class II UIC well. Computes bottomhole flowing pressure from surface "
        "injection pressure accounting for hydrostatic head and tubing friction, then applies Darcy's "
        "steady-state radial flow equation. Yellow cells are inputs; green cells are computed (live formulas).")).font = f_sub
ws.cell(row=2, column=1).fill = PatternFill("solid", fgColor=BRAND)
ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 42

r = 3
# --- well identification (for the submittal) ---
band(ws, r, "Well identification (appears on printed report; not used in calculations)"); r += 1
ident = [
    ("Operator", "wOperator", ""),
    ("Field", "wField", ""),
    ("Well name", "wWell", ""),
    ("API (14-digit)", "wAPI", ""),
    ("UIC code", "wUIC", ""),
    ("Survey date", "wDate", ""),
]
for lbl, nm, val in ident:
    ws.cell(row=r, column=1, value=lbl).font = f_label
    cc = ws.cell(row=r, column=3, value=val); cc.fill = PatternFill("solid", fgColor=HILITE)
    cc.border = box; cc.font = f_input; cc.alignment = left
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    names[nm] = f"Calculator!$C${r}"
    ws.row_dimensions[r].height = 16
    r += 1

# --- 1. Survey data ---
band(ws, r, "1.  Survey data"); r += 1
inp(ws, r, "Surface injection pressure", "P_surf", 500, "psi",
    "Wellhead injection pressure measured during the stabilized portion of the RAT survey. Use the value paired with the observed rate, not a long-term average.",
    name="P_surf"); r += 1
inp(ws, r, "Observed injection rate", "q", 1000, "bbl/d",
    "Instantaneous measured rate at the wellhead flowmeter during the stabilized portion of the survey — NOT the monthly average from injection records. The well should have injected at this rate for 1–2 hours before the reading so pressure has stabilized.",
    name="q"); r += 1
inp(ws, r, "Bottomhole temperature", "T_BH", 120, "°F",
    "Reservoir temperature at the depth of interest. From BHT surveys, the RAT temperature log, or geothermal gradient. Used to auto-compute water viscosity (McCain). Default 120°F is typical for shallow California disposal zones.",
    name="T_BH"); r += 1

# --- 2. Well configuration ---
band(ws, r, "2.  Well configuration"); r += 1
inp(ws, r, "Tubing depth to top perfs", "L", 3000, "ft",
    "Measured depth from surface to the top of the perforated interval (tubing string length). Used for both hydrostatic and friction. For highly deviated wells, hydrostatic should ideally use TVD.",
    name="L"); r += 1
inp(ws, r, "Tubing inside diameter", "d_i", 2.441, "in",
    "Inside diameter of injection tubing. Standard API IDs: 2-3/8\" 4.7# = 1.995; 2-7/8\" 6.5# = 2.441; 3-1/2\" 9.3# = 2.992; 4-1/2\" 12.6# = 3.958 (see Reference Data sheet). Override for worn tubing.",
    name="d_i"); r += 1
inp(ws, r, "Tubing roughness", "ε", 0.0018, "in",
    "Pipe surface roughness. Default 0.0018 in for new commercial steel. Increase to 0.005–0.01 in for scaled, corroded, or service-aged tubing. Friction is moderately sensitive to roughness.",
    name="eps"); r += 1
inp(ws, r, "Wellbore radius", "r_w", 0.354, "ft",
    "Radius of the wellbore at the perforations = (hole diameter in inches) / 24. 8.5-in hole = 0.354 ft; 6.125-in hole = 0.255 ft. Default assumes an 8.5-in hole.",
    name="r_w"); r += 1

# --- 3. Fluid properties ---
band(ws, r, "3.  Fluid properties"); r += 1
inp(ws, r, "Injection fluid gradient", "grad", 0.44, "psi/ft",
    "Hydrostatic gradient of injection fluid. Presets: fresh 0.433; light produced 0.44; typical produced 0.45; heavy brine 0.465 (see Reference Data sheet).",
    name="grad"); r += 1
inp(ws, r, "Fluid density (display)", "ρ", 63.4, "lb/ft³",
    "Density of injection fluid. Converted internally to lb/gal (÷ 7.481) for the Reynolds and Fanning constants. Presets: fresh 62.4; light 63.4; typical 64.9; heavy 67.0.",
    name="rho"); r += 1
inp(ws, r, "Water viscosity (McCain, from T_BH)", "μ_calc", "=109.574*POWER(T_BH,-1.12166)", "cP",
    "Auto-calculated from T_BH via the McCain (1990) fresh-water correlation: μ = 109.574 × T^(-1.12166), T in °F. This is the value used unless an override is entered below.",
    name="mu_calc", is_input=False, numfmt="0.000"); r += 1
inp(ws, r, "Viscosity override (optional)", "μ_ovr", "", "cP",
    "Leave blank to use the McCain value above. Enter a value to override for non-water fluid or a salinity-corrected viscosity.",
    name="mu_ovr", numfmt="0.000"); r += 1
inp(ws, r, "Viscosity used", "μ", "=IF(mu_ovr=\"\",mu_calc,mu_ovr)", "cP",
    "Viscosity carried into the Reynolds, friction, and Darcy calculations: the override if entered, otherwise the McCain value.",
    name="mu", is_input=False, numfmt="0.000"); r += 1
inp(ws, r, "Formation volume factor", "B_w", 1.0, "RB/STB",
    "Water formation volume factor. Default 1.0 is appropriate for typical brines and rarely deviates more than 1–2% for water.",
    name="B_w", numfmt="0.000"); r += 1

# --- 4. Reservoir properties ---
band(ws, r, "4.  Reservoir properties"); r += 1
inp(ws, r, "Static reservoir-pressure mode", "mode", "B", "A / B / C",
    "Selects how static reservoir pressure P_e is determined.  A = direct entry (fall-off extrapolation or known P_e).  B = from static fluid level (default).  C = regional gradient (lowest confidence).",
    name="Pe_mode");
dv = DataValidation(type="list", formula1='"A,B,C"', allow_blank=False); ws.add_data_validation(dv); dv.add(ws.cell(row=r, column=3))
r += 1
inp(ws, r, "Mode A — static reservoir pressure", "P_e(A)", 880, "psi",
    "Used only when mode = A. Reservoir pressure at the drainage radius from a fall-off extrapolation, a direct static gradient survey, or prior work.",
    name="Pe_A"); r += 1
inp(ws, r, "Mode B — static fluid level from surface", "z_fl", 1000, "ft",
    "Used only when mode = B. Depth from surface to the top of the static fluid column under shut-in, from an acoustic fluid-level survey (e.g., Echometer) or wireline static survey. If full to surface, enter 0.",
    name="z_fl"); r += 1
inp(ws, r, "Mode B — static column gradient", "grad_s", 0.44, "psi/ft",
    "Used only when mode = B. Gradient of the static fluid column in the wellbore, typically 0.43–0.46 psi/ft for water/brine.",
    name="grad_s", numfmt="0.000"); r += 1
inp(ws, r, "Mode C — regional pressure gradient", "grad_r", 0.433, "psi/ft",
    "Used only when mode = C. Lowest-confidence option. Typical: 0.40 (lightly depleted), 0.433 (virgin hydrostatic), 0.465 (slightly overpressured).",
    name="grad_r", numfmt="0.000"); r += 1
inp(ws, r, "Drainage radius", "r_e", 660, "ft",
    "Radius at which P_e is assumed to apply. Conventional: 660 ft (40-acre), 933 ft (160-acre), 1320 ft (320-acre). Result is insensitive to r_e because it enters as ln(r_e/r_w).",
    name="r_e"); r += 1
inp(ws, r, "Effective net thickness", "h_eff", 30, "ft",
    "Net thickness actually taking fluid — NOT the total perforated interval. From a spinner survey, sum the intervals showing measurable flow; otherwise use net pay from log analysis.",
    name="h_eff"); r += 1
inp(ws, r, "Perforated interval", "h_perf", 30, "ft",
    "Total perforated interval, for reference only. Not used in k; used to compute fluid-entry efficiency (h_eff / h_perf).",
    name="h_perf"); r += 1
inp(ws, r, "Assumed skin factor", "s", 0, "—",
    "Dimensionless. Default 0 ⇒ result is the effective (skin-included) permeability from the survey. A positive value assumes damage and infers the higher intrinsic k. Single-rate data cannot determine skin — report any non-zero s alongside k.",
    name="s_skin"); r += 1

# --- 5. Intermediate (QC) ---
band(ws, r, "5.  Intermediate values (QC)"); r += 1
def out(ws, row, label, sym, formula, units, helptext, name, numfmt="General"):
    return inp(ws, row, label, sym, formula, units, helptext, name=name, is_input=False, numfmt=numfmt)

out(ws, r, "Density (internal calc)", "ρ_gal", "=rho/7.481", "lb/gal",
    "Display density converted for the Reynolds and Fanning constants: ρ(lb/gal) = ρ(lb/ft³) / 7.481.", "rho_gal", "0.000"); r += 1
out(ws, r, "Reynolds number", "Re", "=11.05*rho_gal*q/(mu*d_i)", "—",
    "Re = 11.05 × ρ(lb/gal) × q / (μ × d_i), q in bbl/d, d_i in inches. Equivalent to the textbook forms 1.48 × ρ(lb/ft³) × q/(d·μ) and 92.1 × SG × q/(d·μ). Flow regime: >4000 turbulent, 2300–4000 transitional, <2300 laminar.", "Re_n", "#,##0"); r += 1
out(ws, r, "Flow regime", "", '=IF(Re_n>4000,"Turbulent",IF(Re_n>2300,"Transitional","Laminar"))', "—",
    "Regime classification implied by Re.", "regime"); r += 1
out(ws, r, "Fanning friction factor", "f", "=IF(Re_n<=4000,16/Re_n,0.0625/(LOG10((eps/d_i)/3.7+5.74/POWER(Re_n,0.9))^2))", "—",
    "Swamee-Jain explicit Fanning factor for Re > 4000; laminar fallback f = 16/Re for Re ≤ 4000.", "f_fan", "0.00000"); r += 1
out(ws, r, "Friction pressure drop", "ΔP_fric", "=5.50E-06*f_fan*rho_gal*q^2*L/POWER(d_i,5)", "psi",
    "Fanning field-unit form: ΔP_f = 5.50×10⁻⁶ × f × ρ(lb/gal) × q² × L / d_i⁵ (f = Fanning friction factor). Equivalent to the textbook v-based form ΔP_f = f × ρ(lb/gal) × v² × L / (25.8 × d_i).", "dP_fric", "0.0"); r += 1
out(ws, r, "Hydrostatic pressure", "P_hydro", "=L*grad", "psi",
    "Pressure added by the fluid column: P_hydro = L × gradient.", "P_hydro", "#,##0"); r += 1
out(ws, r, "Bottomhole flowing pressure", "P_wf", "=P_surf+P_hydro-dP_fric", "psi",
    "P_wf = P_surf + P_hydro − ΔP_fric.", "P_wf", "#,##0"); r += 1
out(ws, r, "Static reservoir pressure (selected)", "P_e", '=IF(Pe_mode="A",Pe_A,IF(Pe_mode="B",(L-z_fl)*grad_s,L*grad_r))', "psi",
    "P_e per the selected mode:  A = direct;  B = (L − z_fl) × grad_s;  C = L × grad_r.", "P_e", "#,##0"); r += 1
out(ws, r, "Mode B validity check", "", '=IF(AND(Pe_mode="B",z_fl>L),"WARNING: fluid level below perfs — Mode B invalid; use A or C","OK")', "—",
    "Flags the Mode B case where the static fluid level lies below the perforations (formation on vacuum / voidage), for which the fluid-level method is not valid.", "modeB_chk"); r += 1
out(ws, r, "Pressure drop across formation", "ΔP", "=P_wf-P_e", "psi",
    "Driving differential for radial flow: ΔP = P_wf − P_e.", "dP_form", "#,##0"); r += 1
out(ws, r, "ln(r_e / r_w)", "", "=LN(r_e/r_w)", "—",
    "Radial geometry term in Darcy's steady-state equation.", "lnRatio", "0.000"); r += 1
out(ws, r, "Fluid entry efficiency", "", "=h_eff/h_perf", "fraction",
    "h_eff / h_perf — fraction of the perforated interval taking fluid.", "entryEff", "0%"); r += 1

# --- 6. Results ---
band(ws, r, "6.  Results"); r += 1
# primary k_eff
ws.cell(row=r, column=1, value="Effective permeability, k_eff").font = f_label
kc = ws.cell(row=r, column=3, value="=141.2*q*mu*B_w*(lnRatio+s_skin)/(h_eff*dP_form)")
kc.font = f_primary; kc.alignment = rightv; kc.border = box
kc.fill = PatternFill("solid", fgColor=RESULT); kc.number_format = "0.0"
names["k_eff"] = f"Calculator!$C${r}"
ws.cell(row=r, column=4, value="mD").font = f_units
ws.cell(row=r, column=5, value="141.2 × q × μ × B × (ln(r_e/r_w) + s) / (h × ΔP), steady-state radial. With s = 0 (default) this is the effective (skin-included) permeability.").font = f_help
ws.cell(row=r, column=5).alignment = leftw
ws.row_dimensions[r].height = 34
r += 1
# numeric-substitution string (live)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
eqc = ws.cell(row=r, column=1, value=('="k = 141.2 × "&TEXT(q,"#,##0")&" × "&TEXT(mu,"0.000")&" × "&TEXT(B_w,"0.0")&" × ("&TEXT(lnRatio,"0.000")&" + "&TEXT(s_skin,"0")&") / ("&TEXT(h_eff,"0")&" × "&TEXT(dP_form,"#,##0")&") = "&TEXT(k_eff,"0.0")&" mD"'))
eqc.font = f_eq; eqc.alignment = left
ws.row_dimensions[r].height = 16
r += 1
out(ws, r, "Permeability-thickness, kh", "kh", "=k_eff*h_eff", "mD-ft",
    "kh = k_eff × h_eff.", "kh", "#,##0"); r += 1
out(ws, r, "Injectivity index, II", "II", "=q/dP_form", "bbl/d/psi",
    "II = q / ΔP.", "II_idx", "0.000"); r += 1
out(ws, r, "Specific injectivity, II/h", "II/h", "=II_idx/h_eff", "bbl/d/psi/ft",
    "II / h_eff.", "IIh", "0.0000"); r += 1

# --- self-check note ---
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
chk = ws.cell(row=r, column=1, value=("Self-check (default inputs): μ = 0.510 cP, Re ≈ 75,200 (turbulent), f = 0.00553, "
    "ΔP_fric = 8.9 psi, P_hydro = 1,320 psi, P_wf = 1,811 psi, P_e = 880 psi, ΔP = 931 psi, "
    "ln(r_e/r_w) = 7.531, k_eff = 19.4 mD, kh = 582 mD-ft, II = 1.074 bbl/d/psi."))
chk.font = Font(name="Calibri", size=8.5, italic=True, color="8A6D00")
chk.fill = PatternFill("solid", fgColor="FFFBEA"); chk.alignment = leftw
ws.row_dimensions[r].height = 28
r += 1

# --- citation footer ---
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
cite = ws.cell(row=r, column=1, value=("Sources: Earlougher (1977), Advances in Well Test Analysis, SPE Monograph 5; "
    "Swamee & Jain (1976), Explicit equations for pipe-flow problems; McCain (1990), The Properties of "
    "Petroleum Fluids, 2nd ed.; Crane Technical Paper No. 410. Field units throughout. See Methods & References sheet."))
cite.font = f_cite; cite.alignment = leftw
ws.row_dimensions[r].height = 26
r += 1

SENS_TITLE_ROW = r + 1
r += 1
# --- 7. Sensitivity ---
band(ws, r, "7.  Sensitivity — k_eff (mD) vs. P_e (columns) and h_eff (rows)"); r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.cell(row=r, column=1, value=("Columns: P_e at ±20% in 10% steps. Rows: h_eff at 0.50× to 1.50× in 25% steps. "
    "All other inputs held at base values; skin held at the entered s. Base case (center) highlighted.")).font = f_help
ws.cell(row=r, column=1).alignment = leftw
ws.row_dimensions[r].height = 26
r += 1

pe_mults = [0.80, 0.90, 1.00, 1.10, 1.20]
h_mults  = [0.50, 0.75, 1.00, 1.25, 1.50]
# header row: corner + Pe multiplier columns (cols B..F => 2..6)
hdr_row = r
corner = ws.cell(row=hdr_row, column=1, value="h_eff ↓  \\  P_e →")
corner.font = f_label; corner.alignment = center; corner.border = box
corner.fill = PatternFill("solid", fgColor=BAND)
mult_row = hdr_row          # store Pe mult in this row, value row below for label
# We'll put Pe multiplier in row hdr_row (hidden numeric) within header text.
for j, pm in enumerate(pe_mults):
    col = 2 + j
    lbl = "base" if pm == 1.0 else (("+" if pm > 1 else "−") + str(int(round(abs(pm-1)*100))) + "%")
    c = ws.cell(row=hdr_row, column=col, value=lbl)
    c.font = f_label; c.alignment = center; c.border = box
    c.fill = PatternFill("solid", fgColor=BAND)
# numeric Pe value sub-row
val_row = hdr_row + 1
ws.cell(row=val_row, column=1, value="P_e (psi) →").font = f_help
ws.cell(row=val_row, column=1).alignment = rightv
ws.cell(row=val_row, column=1).border = box
for j, pm in enumerate(pe_mults):
    col = 2 + j
    c = ws.cell(row=val_row, column=col, value=f"=P_e*{pm}")
    c.font = f_help; c.alignment = center; c.border = box; c.number_format = "#,##0"
# body rows
body_start = val_row + 1
for i, hm in enumerate(h_mults):
    row = body_start + i
    lbl = "base" if hm == 1.0 else (("+" if hm > 1 else "−") + str(int(round(abs(hm-1)*100))) + "%")
    lc = ws.cell(row=row, column=1, value=f'="h {lbl}  ("&TEXT(h_eff*{hm},"0.0")&" ft)"')
    lc.font = f_label; lc.alignment = left; lc.border = box
    lc.fill = PatternFill("solid", fgColor=BAND)
    for j, pm in enumerate(pe_mults):
        col = 2 + j
        # k = 141.2*q*mu*B*(lnRatio+s)/((h_eff*hm)*(P_wf - P_e*pm))
        formula = f"=141.2*q*mu*B_w*(lnRatio+s_skin)/((h_eff*{hm})*(P_wf-P_e*{pm}))"
        c = ws.cell(row=row, column=col, value=formula)
        c.alignment = center; c.border = box; c.number_format = "0.0"
        if hm == 1.0 and pm == 1.0:
            c.font = Font(name="Consolas", size=11, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=ANCHOR)
        else:
            c.font = f_val
r = body_start + len(h_mults) + 1

# ========================================================= Methods sheet =====
wm = wb.create_sheet("Methods & References")
wm.sheet_view.showGridLines = False
wm.column_dimensions["A"].width = 26
wm.column_dimensions["B"].width = 110
wm.merge_cells("A1:B1")
t = wm.cell(row=1, column=1, value="Methods, references, and caveats — Darcy Permeability (RAT)")
t.font = f_title; t.fill = PatternFill("solid", fgColor=BRAND)
t.alignment = Alignment(horizontal="left", vertical="center")
wm.row_dimensions[1].height = 24

methods = [
    ("Darcy radial flow (steady-state, field units)",
     "k (mD) = 141.2 × q (bbl/d) × μ (cP) × B (RB/STB) × (ln(r_e/r_w) + s) / (h (ft) × ΔP (psi)), with ΔP = P_wf − P_e and s = assumed skin (dimensionless; s = 0 reduces to the bare radial form)."),
    ("Bottomhole flowing pressure",
     "P_wf = P_surf + P_hydro − ΔP_fric, with P_hydro = L × gradient."),
    ("Fanning friction equation (field units)",
     "ΔP_f (psi) = 5.50×10⁻⁶ × f × ρ × q² × L / d_i⁵, where f is the Fanning friction factor, ρ in lb/gal (the lb/ft³ input is divided by 7.481), q in bbl/d, L in ft, d_i in inches. Derived from the Fanning relation ΔP = 2·f·ρ·v²·L/D and equal to the v-based field form ΔP_f = f·ρ(lb/gal)·v²·L/(25.8·d_i). (The corresponding lb/ft³ constant is 7.36×10⁻⁷.)"),
    ("Reynolds number",
     "Re = 11.05 × ρ × q / (μ × d_i), ρ in lb/gal, q in bbl/d, d_i in inches, μ in cP. Equivalent to 1.48 × ρ(lb/ft³) × q/(d·μ) and to 92.1 × SG × q/(d·μ), where SG is specific gravity. Use the constant that matches the density unit: pairing the specific-gravity constant 92.1 with lb/gal overstates Re by ~8.3×."),
    ("Swamee-Jain explicit (Fanning)",
     "f = 0.0625 / [log10((ε/d_i)/3.7 + 5.74/Re^0.9)]², valid for Re > 4000. Laminar fallback (Re ≤ 4000): f = 16/Re."),
    ("Water viscosity (McCain 1990, fresh water at atmospheric pressure)",
     "μ (cP) = 109.574 × T (°F)^(-1.12166). Temperature-only fit; no salinity correction in this version. Override the viscosity cell for salinity-corrected or non-water values."),
    ("Density unit convention",
     "Density is entered in lb/ft³ (what petroleum engineers expect) and converted to lb/gal by dividing by 7.481 for the field-unit friction (5.50×10⁻⁶) and Reynolds (11.05) constants used here. The governing requirement is that the constant and the density unit agree: 5.50×10⁻⁶ and 11.05 are the lb/gal forms; 7.36×10⁻⁷ and 1.48 are the equivalent lb/ft³ forms; 92.1 is the specific-gravity Reynolds form. A mismatch (e.g., the SG constant 92.1 applied to lb/gal) scales the result by the density-conversion factor and is the most common error in this calculation."),
    ("Effective vs. intrinsic k",
     "With s = 0 the result is effective (skin-included) permeability. Positive skin (damage) causes the s = 0 result to underestimate intrinsic k; negative skin (stimulation) causes overestimation. To resolve true skin, run a pressure transient analysis. Entering a non-zero s applies the correction explicitly under that assumption — report the assumed s alongside k in any submittal."),
    ("Steady-state assumption",
     "Requires stabilized rate and pressure at the time of the survey. If taken during transient conditions (recent rate change, startup, or shut-in), k_eff will be biased. Allow 1–2 hours of stable injection before recording."),
    ("Observed rate vs. monthly average",
     "Use the wellhead-flowmeter rate during the stabilized portion of the survey — NOT the monthly injection average from regulatory reports. Monthly averages smear over downtime and rate changes and, paired with an instantaneous pressure, give a meaningless result."),
    ("Deviated wells",
     "Friction uses measured depth (correct); hydrostatic also uses measured depth (approximation). For wells > 30° from vertical, hydrostatic should be computed from TVD."),
    ("Single-phase liquid",
     "The friction model assumes single-phase water flow. Multiphase, gassy, or compressible flow requires different methodology."),
    ("Static fluid level (Mode B)",
     "Assumes the static column has equilibrated. For recently shut-in wells the fluid level may still be rising — wait for stabilization before measuring. If the fluid level lies below the perforations, Mode B is invalid (the QC check flags this)."),
    ("Primary references",
     "Earlougher (1977), Advances in Well Test Analysis, SPE Monograph 5; Crane Technical Paper No. 410; Swamee & Jain (1976), “Explicit equations for pipe-flow problems,” J. Hydraulics Div. ASCE; McCain (1990), The Properties of Petroleum Fluids, 2nd ed.; Bourgoyne et al., Applied Drilling Engineering, SPE Textbook Series (density-unit convention)."),
]
mr = 3
for head, body in methods:
    a = wm.cell(row=mr, column=1, value=head); a.font = f_label; a.alignment = leftw
    b = wm.cell(row=mr, column=2, value=body); b.font = Font(name="Calibri", size=10, color="333333"); b.alignment = leftw
    import math
    wm.row_dimensions[mr].height = max(16, math.ceil(len(body)/118)*14 + 6)
    mr += 1

# ========================================================= Reference Data ====
wr = wb.create_sheet("Reference Data")
wr.sheet_view.showGridLines = False
for col, w in {"A": 34, "B": 16, "C": 18, "D": 18, "E": 16}.items():
    wr.column_dimensions[col].width = w
wr.merge_cells("A1:E1")
t = wr.cell(row=1, column=1, value="Reference data and default values")
t.font = f_title; t.fill = PatternFill("solid", fgColor=BRAND); t.alignment = Alignment(horizontal="left", vertical="center")
wr.row_dimensions[1].height = 24

def table(wr, start, title, headers, rows):
    wr.cell(row=start, column=1, value=title).font = f_band
    wr.cell(row=start, column=1).fill = PatternFill("solid", fgColor=BAND)
    wr.merge_cells(start_row=start, start_column=1, end_row=start, end_column=len(headers))
    hr = start + 1
    for j, h in enumerate(headers):
        c = wr.cell(row=hr, column=1+j, value=h); c.font = f_label; c.border = box
        c.alignment = center; c.fill = PatternFill("solid", fgColor="EFF3F5")
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            c = wr.cell(row=hr+1+i, column=1+j, value=v); c.border = box
            c.font = Font(name="Calibri", size=10); c.alignment = left if j == 0 else center
    return hr + 1 + len(rows) + 1

nx = 3
nx = table(wr, nx, "API tubing — OD / weight → inside diameter",
    ["OD / weight", "ID (in)"],
    [["2-3/8\"  4.7 lb/ft", 1.995], ["2-7/8\"  6.5 lb/ft", 2.441],
     ["3-1/2\"  9.3 lb/ft", 2.992], ["4-1/2\"  12.6 lb/ft", 3.958]])
nx = table(wr, nx, "Injection fluid presets",
    ["Fluid type", "Gradient (psi/ft)", "Density (lb/ft³)", "Density (lb/gal)"],
    [["Fresh water", 0.433, 62.4, 8.33], ["Light produced water", 0.44, 63.4, 8.47],
     ["Typical produced water", 0.45, 64.9, 8.67], ["Heavy brine", 0.465, 67.0, 8.96]])
nx = table(wr, nx, "Default reference values",
    ["Parameter", "Default"],
    [["Wellbore radius r_w (8.5-in hole)", "0.354 ft"],
     ["Drainage radius r_e (40-acre)", "660 ft"],
     ["Injection fluid gradient (light produced water)", "0.44 psi/ft"],
     ["Tubing roughness ε (new commercial steel)", "0.0018 in"],
     ["Density conversion", "1 lb/gal = 7.481 lb/ft³"],
     ["Water viscosity correlation", "McCain (1990), temperature-only"]])

# ----------------------------------------------------------- defined names ---
for nm, ref in names.items():
    wb.defined_names.add(DefinedName(name=nm, attr_text=ref))

# recalc on open so Excel/LibreOffice show computed values immediately
wb.calculation.fullCalcOnLoad = True

# print setup: each sheet fits to one page wide
for sheet in (ws, wm, wr):
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = False
    sheet.oddHeader.left.text = "NSLLC Toolkit — Darcy Permeability (RAT)"
    sheet.oddFooter.left.text = "&D"
    sheet.oddFooter.right.text = "Page &P of &N"

out_path = "Darcy-Permeability-RAT.xlsx"
wb.save(out_path)
print("wrote", out_path)
print("defined names:", len(names))
